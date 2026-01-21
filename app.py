from flask import Flask, render_template, request, redirect, url_for, send_from_directory, abort, session, flash, send_file, jsonify
import csv
from datetime import datetime
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'devsecret123')

# Admin credentials (override with env vars in production)
ADMIN_USER = os.environ.get('ADMIN_USER', 'admin')
ADMIN_PASS_HASH = generate_password_hash(os.environ.get('ADMIN_PASS', 'password'))

CSV_FILE = "data/responses.csv"

UPLOAD_PATHS = {
    "invoice": "uploads/invoice",
    "lr": "uploads/lr",
    "discrepancy": "uploads/discrepancy"
}

WAREHOUSE_MAP = {
    "2505": ("Gurugram", "1604"),
    "2507": ("Pune", "1605"),
    "2515": ("Chennai CWH", "1611"),
    "2510": ("Hyderabad", "1607"),
    "2508": ("Chennai Urappakam", "1606"),
    "2517": ("Pillaipakkam", "1612"),
    "2518": ("Coimbatore", "1613"),
    "2512": ("Howrah", "1608")
}

# Ensure folders exist
os.makedirs("data", exist_ok=True)
for path in UPLOAD_PATHS.values():
    os.makedirs(path, exist_ok=True)

# Import parts list from an external Excel (optional). Set PARTS_XLSX_PATH env var to change location.
PARTS_CSV = os.path.join('data','parts.csv')
PARTS_XLSX_PATH = os.environ.get('PARTS_XLSX_PATH', r'c:\\Users\\chill\\OneDrive\\Desktop\\RANE\\part no.xlsx')

def import_parts_from_xlsx(path, out_csv=PARTS_CSV):
    """Import a single-column (or multi-column) XLSX and extract a 'material code' column.
    Writes a CSV with header 'Material Code'."""
    if not os.path.exists(path):
        return False
    try:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            # fallback to pandas if openpyxl not available or fails
            import pandas as pd
            df = pd.read_excel(path, engine='openpyxl')
            rows = [list(df.columns)] + df.fillna('').values.tolist()
    except Exception as e:
        print(f"Failed to read parts xlsx: {e}")
        return False

    if not rows:
        return False
    header = [str(x).strip().lower() if x is not None else '' for x in rows[0]]
    # find a column that looks like material code / part no
    candidates = ['material code','material_code','materialcode','part no','part_no','partno','material','code']
    col_index = None
    for i,h in enumerate(header):
        if any(c in h for c in candidates):
            col_index = i
            break
    # If not found, use first column
    if col_index is None:
        col_index = 0

    values = []
    for r in rows[1:]:
        try:
            v = r[col_index]
        except Exception:
            v = ''
        if v is None:
            continue
        s = str(v).strip()
        if s:
            values.append(s)

    # dedupe and write CSV
    seen = set()
    out_rows = []
    for v in values:
        if v not in seen:
            seen.add(v)
            out_rows.append([v])
    try:
        with open(out_csv, 'w', newline='', encoding='utf-8') as f:
            import csv as _csv
            w = _csv.writer(f)
            w.writerow(['Material Code'])
            w.writerows(out_rows)
        print(f"Imported {len(out_rows)} parts to {out_csv}")
        return True
    except Exception as e:
        print(f"Failed to write parts csv: {e}")
        return False

# Try to import on startup if parts CSV is missing
if not os.path.exists(PARTS_CSV):
    imported = import_parts_from_xlsx(PARTS_XLSX_PATH)
    if not imported:
        print("No parts file found or import failed; continuing without parts list.")

# Company constant
COMPANY_NAME = "RANE LIMITED MADRAS"

# Desired header (includes Reference No)
DESIRED_HEADER = [
    "SR No", "Reference No", "Subject",
    "Invoice No", "Invoice Date",
    "Warehouse", "Warehouse Code",
    "Part No", "Rate",
    "Billed Qty", "Received Qty",
    "Short", "Excess",
    "Mismatch Parts", "Damage",
    "MRP Sticker Required",
    "Customer Name", "City", "State",
    "Transporter Name",
    "LR Number", "LR Date",
    "Invoice Image",
    "LR Image",
    "Discrepancy Image",
    "Status",
    "Admin Notes",
    "Activity Log",
    "Submitted At"
]

# Create CSV if not exists
if not os.path.exists(CSV_FILE):
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(DESIRED_HEADER)
else:
    # Migrate CSV header if Reference No is missing
    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.reader(f))
    if reader:
        header = reader[0]
        # If columns are missing (Reference No, Status, Admin Notes), rebuild rows with new header
        missing = any(col not in header for col in ("Reference No","Status","Admin Notes"))
        if missing:
            new_rows = []
            new_rows.append(DESIRED_HEADER)
            for row in reader[1:]:
                row = list(row)
                if len(row) < 1:
                    continue
                # Start with SR No
                new_row = []
                new_row.append(row[0])  # SR No
                # Reference No (missing) -> blank
                new_row.append('')
                # Remaining old fields (shifted)
                # Fill remaining positions from old row[1:]
                for v in row[1:]:
                    new_row.append(v)
                # Ensure length matches desired header
                while len(new_row) < len(DESIRED_HEADER):
                    new_row.append('')
                # Set customer name (legacy CSVs get default)
                new_row[DESIRED_HEADER.index('Customer Name')] = COMPANY_NAME
                new_rows.append(new_row)
            # write back
            with open(CSV_FILE, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(new_rows)

    else:
        # empty file: write header
        with open(CSV_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(DESIRED_HEADER)


def generate_sr_no():
    with open(CSV_FILE, "r") as f:
        return sum(1 for _ in f)

def get_warehouse(invoice_no):
    if invoice_no and len(invoice_no) >= 4:
        return WAREHOUSE_MAP.get(invoice_no[:4], ("Unknown", "Unknown"))
    return ("Unknown", "Unknown")


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

@app.route('/complain', methods=['GET','POST'])
@app.route('/form', methods=['GET','POST'])
def form():
    sr_no = generate_sr_no()

    if request.method == "POST":

        # Global files (LR and Discrepancy)
        lr_img = request.files.get("lr_image")
        discrepancy_img = request.files.get("discrepancy_image")

        lr_name = ""
        dis_name = ""

        if lr_img and lr_img.filename:
            lr_name = secure_filename(f"{sr_no}_{lr_img.filename}")
            lr_img.save(os.path.join(UPLOAD_PATHS["lr"], lr_name))

        if discrepancy_img and discrepancy_img.filename:
            dis_name = secure_filename(f"{sr_no}_{discrepancy_img.filename}")
            discrepancy_img.save(os.path.join(UPLOAD_PATHS["discrepancy"], dis_name))

        # Add reference number and customer name
        ref_no = f"RLM-{datetime.now().strftime('%Y%m%d')}-{sr_no:05d}"
        customer_name = request.form.get("customer_name") or COMPANY_NAME

        # initial activity log entry
        initial_activity = f"{datetime.now().isoformat()} - Submitted by Portal"

        # Per-invoice fields (lists)
        invoice_nos = request.form.getlist("invoice_no")
        invoice_dates = request.form.getlist("invoice_date")
        part_nos = request.form.getlist("part_no")
        rates = request.form.getlist("rate")
        billed_qtys = request.form.getlist("billed_qty")
        received_qtys = request.form.getlist("received_qty")
        shorts = request.form.getlist("short")
        excesses = request.form.getlist("excess")
        mismatches = request.form.getlist("mismatch")
        damages = request.form.getlist("damage")
        mrp_stickers = request.form.getlist("mrp_sticker")

        invoice_files = request.files.getlist("invoice_image")

        # Determine how many invoice rows to write
        n = max(len(invoice_nos), 1)

        with open(CSV_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            for i in range(n):
                inv_no = invoice_nos[i] if i < len(invoice_nos) else ''
                inv_date = invoice_dates[i] if i < len(invoice_dates) else ''
                warehouse, warehouse_code = get_warehouse(inv_no)
                part_no = part_nos[i] if i < len(part_nos) else ''
                rate = rates[i] if i < len(rates) else ''
                billed_qty = billed_qtys[i] if i < len(billed_qtys) else ''
                received_qty = received_qtys[i] if i < len(received_qtys) else ''
                short = shorts[i] if i < len(shorts) else ''
                excess = excesses[i] if i < len(excesses) else ''
                mismatch = mismatches[i] if i < len(mismatches) else ''
                damage = damages[i] if i < len(damages) else ''
                mrp = mrp_stickers[i] if i < len(mrp_stickers) else ''

                inv_name = ''
                if i < len(invoice_files):
                    fobj = invoice_files[i]
                    if fobj and fobj.filename:
                        inv_name = secure_filename(f"{sr_no}_{i}_{fobj.filename}")
                        fobj.save(os.path.join(UPLOAD_PATHS["invoice"], inv_name))

                row = [
                    sr_no,
                    ref_no,
                    request.form.get("subject"),
                    inv_no,
                    inv_date,
                    warehouse,
                    warehouse_code,
                    part_no,
                    rate,
                    billed_qty,
                    received_qty,
                    short,
                    excess,
                    mismatch,
                    damage,
                    mrp,
                    customer_name,
                    request.form.get("city"),
                    request.form.get("state"),
                    request.form.get("transporter"),
                    request.form.get("lr_no"),
                    request.form.get("lr_date"),
                    inv_name,
                    lr_name,
                    dis_name,
                    '',  # Status
                    '',  # Admin Notes
                    initial_activity,
                    datetime.now().isoformat()
                ]

                writer.writerow(row)

        return redirect(url_for('success', sr_no=sr_no))

    # load parts list if available for datalist/autocomplete
    parts = []
    try:
        if os.path.exists(PARTS_CSV):
            with open(PARTS_CSV, 'r', newline='', encoding='utf-8') as pf:
                preader = list(csv.reader(pf))
                if preader:
                    parts = [r[0] for r in preader[1:] if r]
    except Exception:
        parts = []

    return render_template("form.html", sr_no=sr_no, warehouse_map=WAREHOUSE_MAP, parts=parts)

@app.route('/success/<int:sr_no>')
def success(sr_no):
    # Find the row for this SR No in the CSV and pass some fields to template (use DictReader)
    found = None
    with open(CSV_FILE, 'r', newline='') as f:
        reader = csv.DictReader(f)
        for r in reader:
            try:
                if int(r.get('SR No','0')) == sr_no:
                    found = r
                    break
            except Exception:
                continue
    if not found:
        return render_template('success.html', sr_no=sr_no)

    return render_template('success.html',
                           sr_no=sr_no,
                           ref_no=found.get('Reference No'),
                           subject=found.get('Subject'),
                           invoice_no=found.get('Invoice No'),
                           invoice_img=found.get('Invoice Image'),
                           lr_img=found.get('LR Image'),
                           dis_img=found.get('Discrepancy Image'),
                           submitted_at=found.get('Submitted At'))

@app.route('/uploads/<folder>/<filename>')
def uploaded_file(folder, filename):
    # serve files from uploads subfolders: invoice, lr, discrepancy
    if folder not in UPLOAD_PATHS:
        abort(404)
    directory = os.path.join(app.root_path, UPLOAD_PATHS[folder])
    return send_from_directory(directory, filename)

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == ADMIN_USER and check_password_hash(ADMIN_PASS_HASH, password):
            session['admin_logged_in'] = True
            session['admin_user'] = username
            next_url = request.args.get('next') or url_for('admin')
            return redirect(next_url)
        flash('Invalid credentials', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def home():
    # Landing page where user chooses Customer or Admin
    return render_template('home.html')

@app.route('/admin')
@login_required
def admin():
    q = (request.args.get('q') or '').strip()
    date_from = request.args.get('date_from') or ''
    date_to = request.args.get('date_to') or ''
    warehouse_filter = request.args.get('warehouse') or ''
    part_filter = request.args.get('part_no') or ''
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 10))

    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))

    # full list and warehouses for filter dropdown
    rows_all = reader
    warehouses_all = sorted({(r.get('Warehouse') or 'Unknown') for r in rows_all if r.get('Warehouse')})

    # parts list for filter dropdown: prefer parts CSV if present
    parts_all = []
    try:
        if os.path.exists(PARTS_CSV):
            with open(PARTS_CSV, 'r', newline='', encoding='utf-8') as pf:
                preader = list(csv.reader(pf))
                if preader:
                    parts_all = [r[0] for r in preader[1:] if r]
        else:
            parts_all = sorted({(r.get('Part No') or 'Unknown') for r in rows_all if r.get('Part No')})
    except Exception:
        parts_all = sorted({(r.get('Part No') or 'Unknown') for r in rows_all if r.get('Part No')})

    # start from full list and apply filters
    rows = rows_all

    if q:
        q_lower = q.lower()
        rows = [r for r in rows if q_lower in (r.get('SR No','') or '').lower() or q_lower in (r.get('Reference No','') or '').lower() or q_lower in (r.get('Invoice No','') or '').lower() or q_lower in (r.get('Subject','') or '').lower()]

    if warehouse_filter:
        rows = [r for r in rows if (r.get('Warehouse') or '').lower() == warehouse_filter.lower()]

    if part_filter:
        rows = [r for r in rows if (r.get('Part No') or '').lower() == part_filter.lower()]

    if date_from:
        rows = [r for r in rows if (r.get('Submitted At') or '')[:10] >= date_from]
    if date_to:
        rows = [r for r in rows if (r.get('Submitted At') or '')[:10] <= date_to]

    # show most recent first
    rows = list(reversed(rows))

    total = len(rows)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end = start + per_page
    page_rows = rows[start:end]

    # Group rows by SR No + Reference No so admin sees complaints grouped with multiple invoices
    from collections import OrderedDict
    groups = OrderedDict()
    for r in page_rows:
        key = f"{r.get('SR No','')}|{r.get('Reference No','')}"
        g = groups.setdefault(key, {
            'sr_no': r.get('SR No',''),
            'ref_no': r.get('Reference No',''),
            'subject': r.get('Subject',''),
            'customer': r.get('Customer Name','') or r.get('Company Name',''),
            'submitted_at': r.get('Submitted At',''),
            'status': r.get('Status','') or '',
            'lr_image': r.get('LR Image',''),
            'discrepancy_image': r.get('Discrepancy Image',''),
            'invoices': []
        })
        g['invoices'].append(r)
        # prefer a set status if present
        if not g['status'] and (r.get('Status') or ''):
            g['status'] = r.get('Status')

    return render_template('admin.html', groups=list(groups.values()), q=q, date_from=date_from, date_to=date_to, warehouse_filter=warehouse_filter, warehouses_all=warehouses_all, page=page, pages=pages, per_page=per_page, total=total, parts_all=parts_all, part_filter=part_filter)

@app.route('/admin/view/<int:sr_no>')
@login_required
def admin_view(sr_no):
    rows = []
    with open(CSV_FILE, 'r', newline='') as f:
        reader = csv.DictReader(f)
        for r in reader:
            try:
                if int(r.get('SR No', 0)) == sr_no:
                    rows.append(r)
            except Exception:
                continue
    if not rows:
        return redirect(url_for('admin'))
    # sort invoices by submitted at or leave as-is
    rows = sorted(rows, key=lambda r: r.get('Invoice No') or '')
    return render_template('admin_view.html', rows=rows)

@app.route('/admin/update/<int:sr_no>', methods=['POST'])
@login_required
def admin_update(sr_no):
    status = request.form.get('status') or ''
    notes = request.form.get('notes') or ''
    updated = False
    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))
    # Use DESIRED_HEADER as canonical fieldnames to ensure Activity Log is included
    fieldnames = DESIRED_HEADER
    # Update rows and write back
    new_rows = []
    for r in reader:
        if str(r.get('SR No','')) == str(sr_no):
            r['Status'] = status
            r['Admin Notes'] = notes
            # append to activity log
            existing = r.get('Activity Log') or ''
            new_entry = f"{datetime.now().isoformat()} - {session.get('admin_user','admin')} - {status} - {notes}"
            if existing:
                r['Activity Log'] = existing + '||' + new_entry
            else:
                r['Activity Log'] = new_entry
            updated = True
        new_rows.append(r)
    if updated:
        # sanitize rows to ensure we only write DESIRED_HEADER fields
        sanitized = []
        for row in new_rows:
            clean = {k: (row.get(k) or '') for k in fieldnames}
            sanitized.append(clean)
        with open(CSV_FILE, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(sanitized)
        flash('Record updated', 'info')
    else:
        flash('Record not found', 'error')
    return redirect(url_for('admin_view', sr_no=sr_no))

@app.route('/admin/download')
@login_required
def download_csv():
    # Support download of filtered CSV by accepting same query params as /admin
    q = (request.args.get('q') or '').strip()
    date_from = request.args.get('date_from') or ''
    date_to = request.args.get('date_to') or ''
    warehouse_filter = request.args.get('warehouse') or ''
    part_filter = request.args.get('part_no') or ''

    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))

    rows = reader
    if q:
        q_lower = q.lower()
        rows = [r for r in rows if q_lower in (r.get('SR No','') or '').lower() or q_lower in (r.get('Reference No','') or '').lower() or q_lower in (r.get('Invoice No','') or '').lower() or q_lower in (r.get('Subject','') or '').lower()]
    if warehouse_filter:
        rows = [r for r in rows if (r.get('Warehouse') or '').lower() == warehouse_filter.lower()]
    if part_filter:
        rows = [r for r in rows if (r.get('Part No') or '').lower() == part_filter.lower()]
    if date_from:
        rows = [r for r in rows if (r.get('Submitted At') or '')[:10] >= date_from]
    if date_to:
        rows = [r for r in rows if (r.get('Submitted At') or '')[:10] <= date_to]

    # write to a buffer
    import io

@app.route('/dashboard')
@login_required
def dashboard():
    # collect warehouses for analysis filter
    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))
    warehouses_all = sorted({(r.get('Warehouse') or 'Unknown') for r in reader if r.get('Warehouse')})
    return render_template('dashboard.html', warehouses_all=warehouses_all)


@app.route('/api/parts')
def api_parts():
    # return list of part numbers (Material Code) for autocomplete or admin use
    parts = []
    try:
        if os.path.exists(PARTS_CSV):
            with open(PARTS_CSV, 'r', newline='', encoding='utf-8') as pf:
                preader = list(csv.reader(pf))
                if preader:
                    parts = [r[0] for r in preader[1:] if r]
        else:
            # fallback to scanning existing CSV rows
            with open(CSV_FILE, 'r', newline='') as f:
                reader = list(csv.DictReader(f))
                parts = sorted({(r.get('Part No') or '') for r in reader if r.get('Part No')})
    except Exception:
        parts = []
    return jsonify(parts)

@app.route('/api/stats')
@login_required
def api_stats():
    # return basic stats used by dashboard: timeseries (per date), status counts, top companies
    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))

    from collections import Counter, defaultdict
    times = Counter()
    status = Counter()
    transporters = Counter()
    warehouses = Counter()

    for r in reader:
        submitted = (r.get('Submitted At') or '')[:10]
        if submitted:
            times[submitted] += 1
        s = (r.get('Status') or 'Open') or 'Open'
        status[s] += 1
        t = (r.get('Transporter Name') or 'Unknown')
        transporters[t] += 1
        w = (r.get('Warehouse') or 'Unknown')
        warehouses[w] += 1

    # build timeseries sorted by date
    timeseries = [{'date': d, 'count': times[d]} for d in sorted(times.keys())]
    top_transporters = [{'transporter': t, 'count': transporters[t]} for t in [x[0] for x in transporters.most_common(10)]]

    return {
        'timeseries': timeseries,
        'status': dict(status),
        'top_transporters': top_transporters,
        'warehouses': dict(warehouses)
    }


# ---- Analysis / pivot-like aggregation ----

def _parse_count_value(col_name, raw):
    """Return numeric contribution for a metric column.
    For Short/Excess: parse numeric, default 0.
    For Damage/Mismatch: if numeric use it, else if non-empty treat as 1, else 0.
    """
    if not raw:
        return 0
    raw = str(raw).strip()
    if raw == '':
        return 0
    if col_name in ('Short', 'Excess'):
        try:
            return int(float(raw))
        except Exception:
            return 0
    else:
        # Damage / Mismatch -> count occurrences
        try:
            return int(float(raw))
        except Exception:
            return 1


def compute_analysis(rows, metric='short', group_by='transporter', date_from='', date_to='', top_n=10, warehouse_filter=''):
    """Compute aggregated values for metric grouped by group_by.
    - metric: 'short','excess','damage','mismatch'
    - group_by: 'transporter','company','warehouse','warehouse_code','invoice','part_no','city'
    """
    metric = (metric or 'short').lower()
    col_map = {'short': 'Short', 'excess': 'Excess', 'damage': 'Damage', 'mismatch': 'Mismatch Parts'}
    gb_map = {'transporter': 'Transporter Name', 'company': 'Customer Name', 'customer': 'Customer Name', 'warehouse': 'Warehouse', 'warehouse_code': 'Warehouse Code', 'invoice': 'Invoice No', 'part_no': 'Part No', 'city': 'City'}

    col = col_map.get(metric, 'Short')
    gb_col = gb_map.get(group_by, group_by)

    counts = {}
    for r in rows:
        submitted = (r.get('Submitted At') or '')[:10]
        if date_from and submitted and submitted < date_from:
            continue
        if date_to and submitted and submitted > date_to:
            continue
        if warehouse_filter:
            if (r.get('Warehouse') or '').lower() != warehouse_filter.lower():
                continue
        group = (r.get(gb_col) or 'Unknown')
        value = _parse_count_value(col, r.get(col))
        counts[group] = counts.get(group, 0) + value

    items = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    if top_n:
        items = items[:int(top_n)]
    return [{'group': k, 'value': v} for k, v in items]


@app.route('/api/analysis/export')
@login_required
def api_analysis_export():
    metric = (request.args.get('metric') or 'short').lower()
    group_by = (request.args.get('group_by') or 'transporter').lower()
    top_n = int(request.args.get('top_n') or 0)
    date_from = request.args.get('date_from') or ''
    date_to = request.args.get('date_to') or ''
    warehouse_filter = request.args.get('warehouse') or ''
    fmt = (request.args.get('format') or 'csv').lower()

    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))

    rows = compute_analysis(reader, metric=metric, group_by=group_by, date_from=date_from, date_to=date_to, top_n=top_n or None, warehouse_filter=warehouse_filter)

    # CSV fallback
    import io
    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['Group', 'Value'])
            for r in rows:
                ws.append([r['group'], r['value']])
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'analysis_{metric}_{group_by}.xlsx')
        except Exception:
            # fallback to csv
            fmt = 'csv'
    # CSV
    if fmt == 'csv':
        out = io.StringIO()
        out.write('Group,Value\n')
        for r in rows:
            # escape quotes
            g = str(r['group']).replace('"','""')
            out.write(f'"{g}",{r["value"]}\n')
        out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name=f'analysis_{metric}_{group_by}.csv')

    # default
    return {'metric': metric, 'group_by': group_by, 'rows': rows}

# Simple in-memory cache for analysis results
ANALYSIS_CACHE = {}
ANALYSIS_TTL = 30  # seconds


def get_cached_analysis(metric, group_by, top_n, date_from, date_to, warehouse_filter):
    from time import time
    key = (metric, group_by, int(top_n) if top_n else None, date_from or '', date_to or '', warehouse_filter or '')
    now = time()
    entry = ANALYSIS_CACHE.get(key)
    if entry:
        ts, value = entry
        if now - ts < ANALYSIS_TTL:
            return value
    # compute fresh
    with open(CSV_FILE, 'r', newline='') as f:
        reader = list(csv.DictReader(f))
    value = compute_analysis(reader, metric=metric, group_by=group_by, date_from=date_from, date_to=date_to, top_n=top_n, warehouse_filter=warehouse_filter)
    ANALYSIS_CACHE[key] = (now, value)
    return value


@app.route('/api/analysis')
@login_required
def api_analysis():
    metric = (request.args.get('metric') or 'short').lower()
    group_by = (request.args.get('group_by') or 'transporter').lower()
    top_n = int(request.args.get('top_n') or 10)
    date_from = request.args.get('date_from') or ''
    date_to = request.args.get('date_to') or ''
    warehouse_filter = request.args.get('warehouse') or ''

    rows = get_cached_analysis(metric, group_by, top_n, date_from, date_to, warehouse_filter)
    return {'metric': metric, 'group_by': group_by, 'rows': rows}

@app.route('/favicon.ico')
def favicon():
    # Redirect legacy /favicon.ico requests to the SVG favicon in static
    return redirect(url_for('static', filename='favicon.svg'))


@app.route('/track', methods=['GET','POST'])
def track():
    result = None
    ref = ''
    if request.method == 'POST':
        ref = (request.form.get('reference') or '').strip()
        if ref:
            with open(CSV_FILE, 'r', newline='') as f:
                reader = csv.DictReader(f)
                for r in reader:
                    if (r.get('Reference No') or '').strip() == ref:
                        result = r
                        break
    return render_template('track.html', result=result, ref=ref)

@app.route('/track/<ref_no>')
def track_direct(ref_no):
    result = None
    with open(CSV_FILE, 'r', newline='') as f:
        reader = csv.DictReader(f)
        for r in reader:
            if (r.get('Reference No') or '').strip() == ref_no:
                result = r
                break
    return render_template('track.html', result=result, ref=ref_no)

if __name__ == "__main__":
    app.run(debug=True)

